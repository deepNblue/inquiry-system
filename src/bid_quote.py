"""
投标报价模块
- 分项报价清单（标准投标格式）
- 投标总价汇总（含税/不含税/税率）
- 技术偏离表生成
- 商务偏离表生成
"""

from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass, field
from enum import Enum
from datetime import datetime


class DeviationType(Enum):
    """偏离类型"""
    POSITIVE = "正偏离"      # 我方优于招标文件
    NEGATIVE = "负偏离"      # 我方低于招标文件
    NONE = "无偏离"          # 完全响应


@dataclass
class QuoteItem:
    """报价单项"""
    seq: int                    # 序号
    name: str                   # 设备名称
    brand: str                  # 品牌
    model: str                  # 型号规格
    unit: str = "套"           # 单位
    quantity: int = 1          # 数量
    unit_price: float = 0       # 单价
    tax_rate: float = 0.13      # 税率
    origin: str = "中国"        # 产地
    delivery: str = "30天"      # 交货期
    warranty: str = "3年"      # 质保期
    deviation: str = "无偏离"   # 偏离情况
    deviation_note: str = ""    # 偏离说明
    
    @property
    def tax_price(self) -> float:
        """含税单价"""
        return self.unit_price * (1 + self.tax_rate)
    
    @property
    def subtotal(self) -> float:
        """合价（不含税）"""
        return self.unit_price * self.quantity
    
    @property
    def tax_subtotal(self) -> float:
        """合价（含税）"""
        return self.tax_price * self.quantity


@dataclass
class BidQuoteConfig:
    """投标报价配置"""
    project_name: str = ""              # 项目名称
    bid_number: str = ""                # 招标编号
    bidder_name: str = ""               # 投标人名称
    tax_rate: float = 0.13              # 默认税率
    delivery_default: str = "30天"      # 默认交货期
    warranty_default: str = "3年"       # 默认质保期
    valid_days: int = 90               # 报价有效期


class BidQuoteGenerator:
    """投标报价生成器"""
    
    def __init__(self, config: BidQuoteConfig = None):
        self.config = config or BidQuoteConfig()
        self.items: List[QuoteItem] = []
    
    def add_item(self, item: QuoteItem):
        """添加报价项"""
        if item.seq == 0:
            item.seq = len(self.items) + 1
        self.items.append(item)
    
    def add_from_result(self, result: Dict, seq: int = 0) -> QuoteItem:
        """从询价结果添加"""
        item = QuoteItem(
            seq=seq or len(self.items) + 1,
            name=result.get("name", ""),
            brand=result.get("brand", "待定"),
            model=result.get("model", result.get("specs", "")),
            unit_price=result.get("price", 0),
            unit=result.get("unit", "套"),
            quantity=result.get("quantity", 1),
            warranty=self.config.warranty_default,
            delivery=self.config.delivery_default,
        )
        self.add_item(item)
        return item
    
    def calculate(self) -> Dict:
        """计算汇总"""
        subtotal_no_tax = sum(i.subtotal for i in self.items)
        tax_total = sum(i.tax_subtotal for i in self.items) - subtotal_no_tax
        total_with_tax = subtotal_no_tax + tax_total
        
        return {
            "subtotal_no_tax": subtotal_no_tax,
            "tax_amount": tax_total,
            "total_with_tax": total_with_tax,
            "tax_rate": self.config.tax_rate,
            "item_count": len(self.items),
            "quantity_total": sum(i.quantity for i in self.items),
        }
    
    def generate_quote_table(self) -> str:
        """生成分项报价清单"""
        calc = self.calculate()
        
        lines = [
            f"# 分项报价清单",
            f"",
            f"**项目名称**: {self.config.project_name}",
            f"**招标编号**: {self.config.bid_number}",
            f"**投标人**: {self.config.bidder_name}",
            f"**报价日期**: {datetime.now().strftime('%Y年%m月%d日')}",
            f"",
            f"---",
            f"",
            f"## 报价明细",
            f"",
            f"| 序号 | 设备名称 | 品牌 | 型号规格 | 单位 | 数量 | 单价 | 合价 | 税率 | 交货期 | 质保 | 偏离 |",
            f"|------|----------|------|----------|------|------|------|------|------|--------|------|------|",
        ]
        
        for item in self.items:
            lines.append(
                f"| {item.seq} | {item.name} | {item.brand} | {item.model[:30]} | "
                f"{item.unit} | {item.quantity} | ¥{item.unit_price:,.2f} | "
                f"¥{item.subtotal:,.2f} | {item.tax_rate*100:.0f}% | "
                f"{item.delivery} | {item.warranty} | {item.deviation} |"
            )
        
        lines.extend([
            f"",
            f"---",
            f"",
            f"## 报价汇总",
            f"",
            f"| 项目 | 金额 |",
            f"|------|------|",
            f"| 合价（不含税） | ¥{calc['subtotal_no_tax']:,.2f} |",
            f"| 税额（{calc['tax_rate']*100:.0f}%） | ¥{calc['tax_amount']:,.2f} |",
            f"| **合价（含税）** | **¥{calc['total_with_tax']:,.2f}** |",
            f"",
            f"---",
            f"",
            f"**报价有效期**: {self.config.valid_days}天",
            f"",
        ])
        
        return "\n".join(lines)
    
    def generate_deviation_table(self, specs_requirements: List[Dict]) -> str:
        """生成技术偏离表"""
        lines = [
            "# 技术偏离表",
            "",
            f"**项目名称**: {self.config.project_name}",
            f"**招标编号**: {self.config.bid_number}",
            "",
            f"---",
            "",
            f"## 一、技术参数偏离",
            "",
            f"| 序号 | 设备名称 | 招标文件要求 | 我方响应 | 偏离情况 | 说明 |",
            f"|------|----------|-------------|----------|----------|------|",
        ]
        
        seq = 1
        for req in specs_requirements:
            item_name = req.get("name", "")
            spec_required = req.get("spec_required", "")
            spec_response = req.get("spec_response", "")
            deviation = req.get("deviation", "无偏离")
            note = req.get("note", "")
            
            # 简化规格显示
            spec_required = spec_required[:50] if len(spec_required) > 50 else spec_required
            spec_response = spec_response[:50] if len(spec_response) > 50 else spec_response
            
            lines.append(
                f"| {seq} | {item_name} | {spec_required} | "
                f"{spec_response} | {deviation} | {note} |"
            )
            seq += 1
        
        lines.extend([
            "",
            f"---",
            "",
            f"## 二、商务偏离",
            "",
            f"| 序号 | 条款 | 招标文件要求 | 我方响应 | 偏离 |",
            f"|------|------|-------------|----------|------|",
            f"| 1 | 交货期 | 合同签订后30天内 | {self.config.delivery_default} | 无偏离 |",
            f"| 2 | 质保期 | 不少于2年 | {self.config.warranty_default} | 正偏离 |",
            f"| 3 | 付款方式 | 预付30%+交货70% | 预付30%+验收30%+交货40% | 无偏离 |",
            f"| 4 | 培训 | 免费培训 | 免费培训 | 无偏离 |",
            f"| 5 | 售后服务 | 7×24小时响应 | 7×24小时响应 | 无偏离 |",
            "",
        ])
        
        return "\n".join(lines)
    
    def generate_full_document(self, specs_requirements: List[Dict] = None) -> str:
        """生成完整投标文件（简化版）"""
        sections = [
            self.generate_quote_table(),
        ]
        
        if specs_requirements:
            sections.append(self.generate_deviation_table(specs_requirements))
        
        # 汇总信息
        calc = self.calculate()
        sections.append(f"""
---

# 投标总览

## 价格汇总

| 项目 | 金额 |
|------|------|
| 投标报价（不含税） | ¥{calc['subtotal_no_tax']:,.2f} |
| 税率（{calc['tax_rate']*100:.0f}%） | ¥{calc['tax_amount']:,.2f} |
| **投标报价（含税）** | **¥{calc['total_with_tax']:,.2f}** |

## 产品汇总

- 设备数量: {calc['quantity_total']} 台/套
- 报价项数: {calc['item_count']} 项
- 报价有效期: {self.config.valid_days}天

## 品牌构成

""")
        
        # 品牌汇总
        brands = {}
        for item in self.items:
            if item.brand not in brands:
                brands[item.brand] = {"count": 0, "amount": 0}
            brands[item.brand]["count"] += 1
            brands[item.brand]["amount"] += item.subtotal
        
        sections.append("| 品牌 | 数量 | 金额 |")
        sections.append("|------|------|------|")
        for brand, data in sorted(brands.items(), key=lambda x: x[1]["amount"], reverse=True):
            sections.append(f"| {brand} | {data['count']}项 | ¥{data['amount']:,.2f} |")
        
        return "\n".join(sections)
    
    def export_excel(self, output_path: str):
        """导出Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "报价清单"
            
            # 表头
            headers = ["序号", "设备名称", "品牌", "型号规格", "单位", "数量", 
                      "单价(不含税)", "合价(不含税)", "税率", "交货期", "质保", "偏离"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # 数据
            for row, item in enumerate(self.items, 2):
                ws.cell(row=row, column=1, value=item.seq)
                ws.cell(row=row, column=2, value=item.name)
                ws.cell(row=row, column=3, value=item.brand)
                ws.cell(row=row, column=4, value=item.model)
                ws.cell(row=row, column=5, value=item.unit)
                ws.cell(row=row, column=6, value=item.quantity)
                ws.cell(row=row, column=7, value=item.unit_price)
                ws.cell(row=row, column=8, value=item.subtotal)
                ws.cell(row=row, column=9, value=f"{item.tax_rate*100:.0f}%")
                ws.cell(row=row, column=10, value=item.delivery)
                ws.cell(row=row, column=11, value=item.warranty)
                ws.cell(row=row, column=12, value=item.deviation)
            
            # 汇总行
            calc = self.calculate()
            last_row = len(self.items) + 2
            ws.cell(row=last_row, column=7, value="合计")
            ws.cell(row=last_row, column=8, value=calc["subtotal_no_tax"])
            
            # 设置列宽
            widths = [8, 25, 15, 35, 8, 8, 15, 15, 8, 12, 10, 10]
            for i, width in enumerate(widths, 1):
                ws.column_dimensions[chr(64+i)].width = width
            
            wb.save(output_path)
            return output_path
            
        except ImportError:
            print("⚠️ 需要安装 openpyxl: pip install openpyxl")
            return None


def compare_specs(spec_required: str, spec_response: str) -> Tuple[DeviationType, str]:
    """
    比较规格参数，返回偏离类型和说明
    """
    req_lower = spec_required.lower()
    res_lower = spec_response.lower()
    
    # 完全匹配
    if req_lower == res_lower:
        return DeviationType.NONE, "完全响应"
    
    # 检查关键参数
    import re
    
    # 提取数字参数
    def extract_numbers(text):
        return [float(n) for n in re.findall(r'\d+\.?\d*', text)]
    
    req_nums = extract_numbers(spec_required)
    res_nums = extract_numbers(spec_response)
    
    if req_nums and res_nums:
        # 数值比较
        res_num = res_nums[0]
        req_num = req_nums[0]
        
        if res_num > req_num:
            return DeviationType.POSITIVE, f"优于要求({res_num}>{req_num})"
        elif res_num < req_num:
            return DeviationType.NEGATIVE, f"低于要求({res_num}<{req_num})"
    
    return DeviationType.NONE, "基本响应"
