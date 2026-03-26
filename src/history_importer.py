"""
历史数据导入模块
支持 Excel、CSV、JSON 等多种格式
"""

import os
import json
import csv
from typing import List, Dict, Optional
from datetime import datetime
from dataclasses import dataclass


@dataclass
class ImportResult:
    """导入结果"""
    total_rows: int
    success_count: int
    failed_count: int
    errors: List[str]
    imported_records: List[Dict]


class HistoryImporter:
    """
    历史数据导入器
    支持多种文件格式
    """
    
    # 支持的文件格式
    SUPPORTED_FORMATS = ['.csv', '.xlsx', '.xls', '.json', '.txt']
    
    def __init__(self):
        self.errors = []
    
    def import_file(self, file_path: str, format_hint: str = None) -> ImportResult:
        """
        导入文件
        
        Args:
            file_path: 文件路径
            format_hint: 格式提示 (csv/excel/json)
        
        Returns:
            导入结果
        """
        if not os.path.exists(file_path):
            return ImportResult(
                total_rows=0,
                success_count=0,
                failed_count=0,
                errors=[f"文件不存在: {file_path}"],
                imported_records=[]
            )
        
        # 自动检测格式
        ext = format_hint or os.path.splitext(file_path)[1].lower()
        
        try:
            if ext == '.csv' or format_hint == 'csv':
                return self._import_csv(file_path)
            elif ext in ['.xlsx', '.xls'] or format_hint == 'excel':
                return self._import_excel(file_path)
            elif ext == '.json' or format_hint == 'json':
                return self._import_json(file_path)
            elif ext == '.txt':
                return self._import_txt(file_path)
            else:
                return ImportResult(
                    total_rows=0,
                    success_count=0,
                    failed_count=0,
                    errors=[f"不支持的格式: {ext}"],
                    imported_records=[]
                )
        except Exception as e:
            return ImportResult(
                total_rows=0,
                success_count=0,
                failed_count=0,
                errors=[f"导入失败: {str(e)}"],
                imported_records=[]
            )
    
    def _import_csv(self, file_path: str) -> ImportResult:
        """导入 CSV 文件"""
        records = []
        errors = []
        success = 0
        failed = 0
        
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            # 尝试检测分隔符
            sample = f.read(1024)
            f.seek(0)
            
            delimiter = ',' if ',' in sample else '\t'
            
            reader = csv.DictReader(f, delimiter=delimiter)
            
            for i, row in enumerate(reader, 1):
                try:
                    record = self._parse_record(row)
                    if record:
                        records.append(record)
                        success += 1
                    else:
                        failed += 1
                        errors.append(f"行 {i}: 数据不完整")
                except Exception as e:
                    failed += 1
                    errors.append(f"行 {i}: {str(e)}")
        
        return ImportResult(
            total_rows=success + failed,
            success_count=success,
            failed_count=failed,
            errors=errors,
            imported_records=records
        )
    
    def _import_excel(self, file_path: str) -> ImportResult:
        """导入 Excel 文件"""
        try:
            import openpyxl
        except ImportError:
            return ImportResult(
                total_rows=0,
                success_count=0,
                failed_count=0,
                errors=["需要安装 openpyxl: pip install openpyxl"],
                imported_records=[]
            )
        
        records = []
        errors = []
        success = 0
        failed = 0
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 获取表头
        headers = [cell.value for cell in ws[1]]
        headers = [h.strip() if h else f"col_{i}" for i, h in enumerate(headers)]
        
        # 映射常见列名
        col_map = self._map_columns(headers)
        
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                row_dict = {headers[j]: row[j] for j in range(len(headers))}
                record = self._parse_record(row_dict, col_map)
                
                if record:
                    records.append(record)
                    success += 1
                else:
                    failed += 1
                    errors.append(f"行 {i}: 数据不完整")
            except Exception as e:
                failed += 1
                errors.append(f"行 {i}: {str(e)}")
        
        wb.close()
        
        return ImportResult(
            total_rows=success + failed,
            success_count=success,
            failed_count=failed,
            errors=errors[:100],  # 限制错误数量
            imported_records=records
        )
    
    def _import_json(self, file_path: str) -> ImportResult:
        """导入 JSON 文件"""
        records = []
        errors = []
        
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 支持数组或对象
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            # 尝试找数据数组
            items = data.get('data', data.get('records', data.get('items', [data])))
        else:
            return ImportResult(
                total_rows=0,
                success_count=0,
                failed_count=0,
                errors=["JSON 格式不正确"],
                imported_records=[]
            )
        
        for i, item in enumerate(items, 1):
            try:
                record = self._parse_record(item)
                if record:
                    records.append(record)
            except Exception as e:
                errors.append(f"项 {i}: {str(e)}")
        
        return ImportResult(
            total_rows=len(items),
            success_count=len(records),
            failed_count=len(items) - len(records),
            errors=errors,
            imported_records=records
        )
    
    def _import_txt(self, file_path: str) -> ImportResult:
        """导入 TXT 文件 (每行一个产品)"""
        records = []
        errors = []
        
        with open(file_path, 'r', encoding='utf-8') as f:
            for i, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                
                # 尝试解析为 JSON
                if line.startswith('{'):
                    try:
                        item = json.loads(line)
                        record = self._parse_record(item)
                        if record:
                            records.append(record)
                    except:
                        errors.append(f"行 {i}: JSON 解析失败")
                else:
                    # 纯文本作为产品名
                    records.append({
                        'product_name': line,
                        'timestamp': datetime.now().isoformat(),
                    })
        
        return ImportResult(
            total_rows=len(records) + len(errors),
            success_count=len(records),
            failed_count=len(errors),
            errors=errors,
            imported_records=records
        )
    
    def _map_columns(self, headers: List[str]) -> Dict[str, str]:
        """映射列名"""
        col_map = {}
        
        # 常见列名映射
        mappings = {
            'product_name': ['产品名称', '产品名', '名称', 'name', '产品', '设备名称', '设备名'],
            'brand': ['品牌', '厂家', 'manufacturer', 'make'],
            'model': ['型号', 'model', '规格型号', '货号'],
            'specs': ['规格', '参数', '技术参数', 'specs', '规格参数'],
            'price': ['价格', '单价', '报价', 'price', '含税价'],
            'quantity': ['数量', 'qty', 'num', 'count'],
            'unit': ['单位', 'unit'],
            'source': ['来源', '供应商', '渠道', 'source'],
            'date': ['日期', '时间', 'date', 'timestamp'],
        }
        
        for standard, aliases in mappings.items():
            for alias in aliases:
                for header in headers:
                    if alias.lower() in header.lower():
                        col_map[standard] = header
                        break
        
        return col_map
    
    def _parse_record(self, data: Dict, col_map: Dict = None) -> Optional[Dict]:
        """解析单条记录"""
        if not data:
            return None
        
        record = {}
        
        # 产品名称 (必填)
        name = data.get('product_name') or data.get('name') or data.get('产品名称')
        if not name:
            # 尝试通过 col_map
            if col_map and 'product_name' in col_map:
                name = data.get(col_map['product_name'])
        
        if not name:
            return None
        
        record['product_name'] = str(name).strip()
        
        # 品牌
        record['brand'] = str(data.get('brand', data.get('品牌', ''))).strip()
        
        # 型号
        record['model'] = str(data.get('model', data.get('型号', ''))).strip()
        
        # 规格
        record['specs'] = str(data.get('specs', data.get('规格', ''))).strip()
        
        # 价格
        price = data.get('price', data.get('价格', 0))
        try:
            record['price'] = float(price) if price else 0
        except (ValueError, TypeError):
            record['price'] = 0
        
        # 数量
        qty = data.get('quantity', data.get('数量', 1))
        try:
            record['quantity'] = int(qty) if qty else 1
        except (ValueError, TypeError):
            record['quantity'] = 1
        
        # 单位
        record['unit'] = str(data.get('unit', data.get('单位', '台'))).strip()
        
        # 来源
        record['source'] = str(data.get('source', data.get('来源', '历史导入'))).strip()
        
        # 时间戳
        record['timestamp'] = data.get('date') or data.get('timestamp') or datetime.now().isoformat()
        
        return record
    
    def save_to_database(self, records: List[Dict], db_path: str = "data/history.db") -> bool:
        """保存到数据库"""
        if not records:
            return False
        
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        
        try:
            import sqlite3
            
            conn = sqlite3.connect(db_path)
            
            # 创建表
            conn.execute("""
                CREATE TABLE IF NOT EXISTS price_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_name TEXT NOT NULL,
                    brand TEXT,
                    model TEXT,
                    specs TEXT,
                    price REAL,
                    quantity INTEGER DEFAULT 1,
                    unit TEXT DEFAULT '台',
                    source TEXT,
                    timestamp TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # 插入数据
            for r in records:
                conn.execute("""
                    INSERT INTO price_history 
                    (product_name, brand, model, specs, price, quantity, unit, source, timestamp)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    r.get('product_name'),
                    r.get('brand'),
                    r.get('model'),
                    r.get('specs'),
                    r.get('price', 0),
                    r.get('quantity', 1),
                    r.get('unit', '台'),
                    r.get('source', '导入'),
                    r.get('timestamp', datetime.now().isoformat())
                ))
            
            conn.commit()
            conn.close()
            
            print(f"✓ 已保存 {len(records)} 条记录到数据库")
            return True
            
        except Exception as e:
            print(f"✗ 保存失败: {e}")
            return False


# 便捷函数
def import_history(file_path: str, format_hint: str = None) -> ImportResult:
    """快速导入"""
    importer = HistoryImporter()
    return importer.import_file(file_path, format_hint)
