"""
数据导出模块
支持多种格式的询价结果导出
"""

import os
import json
import csv
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path


class DataExporter:
    """
    数据导出器
    支持 Markdown, JSON, CSV, Excel, HTML 格式
    """
    
    def __init__(self, output_dir: str = "output"):
        self.output_dir = output_dir
        Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    def export(
        self,
        results: List[Any],
        filename: str,
        format: str = "markdown"
    ) -> str:
        """
        导出数据
        
        Args:
            results: 询价结果列表
            filename: 文件名（不含扩展名）
            format: 格式 (markdown/json/csv/html/excel)
        
        Returns:
            输出文件路径
        """
        os.makedirs(self.output_dir, exist_ok=True)
        
        if format == "markdown":
            return self._export_markdown(results, filename)
        elif format == "json":
            return self._export_json(results, filename)
        elif format == "csv":
            return self._export_csv(results, filename)
        elif format == "html":
            return self._export_html(results, filename)
        elif format == "excel":
            return self._export_excel(results, filename)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _export_markdown(self, results: List[Any], filename: str) -> str:
        """导出 Markdown"""
        from src.aggregator import PriceAggregator
        
        aggregator = PriceAggregator()
        content = aggregator.generate_report(results, "markdown")
        
        path = f"{self.output_dir}/{filename}.md"
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        
        return path
    
    def _export_json(self, results: List[Any], filename: str) -> str:
        """导出 JSON"""
        data = []
        for r in results:
            if hasattr(r, "to_dict"):
                data.append(r.to_dict())
            elif hasattr(r, "__dict__"):
                data.append(self._clean_dict(r.__dict__))
        
        content = json.dumps(data, ensure_ascii=False, indent=2)
        
        path = f"{self.output_dir}/{filename}.json"
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        
        return path
    
    def _export_csv(self, results: List[Any], filename: str) -> str:
        """导出 CSV"""
        path = f"{self.output_dir}/{filename}.csv"
        
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            
            # 表头
            writer.writerow(["产品名称", "品牌", "型号", "最低价", "最高价", "平均价", "推荐来源", "推荐价格", "报价数量", "更新时间"])
            
            # 数据
            for r in results:
                writer.writerow([
                    getattr(r, "product_name", ""),
                    getattr(r, "brand", ""),
                    getattr(r, "model", ""),
                    getattr(r, "min_price", 0),
                    getattr(r, "max_price", 0),
                    getattr(r, "avg_price", 0),
                    getattr(r, "recommended_source", ""),
                    getattr(r, "recommended_price", 0),
                    getattr(r, "source_count", 0),
                    getattr(r, "last_updated", ""),
                ])
        
        return path
    
    def _export_html(self, results: List[Any], filename: str) -> str:
        """导出 HTML"""
        html = self._generate_html(results)
        
        path = f"{self.output_dir}/{filename}.html"
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        
        return path
    
    def _export_excel(self, results: List[Any], filename: str) -> str:
        """导出 Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("openpyxl required: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "询价报告"
        
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        headers = ["产品名称", "品牌", "型号", "最低价", "最高价", "平均价", "推荐来源", "推荐价格", "报价数量"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # 写入数据
        for row, r in enumerate(results, 2):
            ws.cell(row=row, column=1, value=getattr(r, "product_name", ""))
            ws.cell(row=row, column=2, value=getattr(r, "brand", ""))
            ws.cell(row=row, column=3, value=getattr(r, "model", ""))
            ws.cell(row=row, column=4, value=getattr(r, "min_price", 0))
            ws.cell(row=row, column=5, value=getattr(r, "max_price", 0))
            ws.cell(row=row, column=6, value=getattr(r, "avg_price", 0))
            ws.cell(row=row, column=7, value=getattr(r, "recommended_source", ""))
            ws.cell(row=row, column=8, value=getattr(r, "recommended_price", 0))
            ws.cell(row=row, column=9, value=getattr(r, "source_count", 0))
        
        # 调整列宽
        for col in range(1, 10):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        path = f"{self.output_dir}/{filename}.xlsx"
        wb.save(path)
        
        return path
    
    def _generate_html(self, results: List[Any]) -> str:
        """生成 HTML"""
        rows = []
        for r in results:
            min_p = getattr(r, "min_price", 0)
            max_p = getattr(r, "max_price", 0)
            
            if min_p > 0:
                price_str = f"¥{min_p:,.0f} ~ ¥{max_p:,.0f}"
            else:
                price_str = "待报价"
            
            rows.append(f"""
            <tr>
                <td>{getattr(r, 'product_name', '')}</td>
                <td>{getattr(r, 'brand', '')}</td>
                <td>{getattr(r, 'recommended_source', '')}</td>
                <td>{price_str}</td>
                <td>{getattr(r, 'source_count', 0)}</td>
            </tr>
            """)
        
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>询价报告</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #333; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
        th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
        th {{ background-color: #366092; color: white; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
        .price {{ color: #e74c3c; font-weight: bold; }}
        .meta {{ color: #666; font-size: 14px; margin-top: 20px; }}
    </style>
</head>
<body>
    <h1>📊 询价报告</h1>
    <div class="meta">
        <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p>产品数量: {len(results)}</p>
    </div>
    
    <table>
        <thead>
            <tr>
                <th>产品名称</th>
                <th>品牌</th>
                <th>推荐来源</th>
                <th>价格区间</th>
                <th>报价数量</th>
            </tr>
        </thead>
        <tbody>
            {''.join(rows)}
        </tbody>
    </table>
</body>
</html>
        """
        
        return html
    
    def _clean_dict(self, d: Dict) -> Dict:
        """清理字典（移除不可序列化对象）"""
        result = {}
        for k, v in d.items():
            if isinstance(v, (str, int, float, bool, type(None))):
                result[k] = v
            elif isinstance(v, list):
                result[k] = [self._clean_dict(item) if isinstance(item, dict) else str(item) for item in v]
            elif isinstance(v, dict):
                result[k] = self._clean_dict(v)
            else:
                result[k] = str(v)
        return result


# 便捷函数
def export_results(results: List, filename: str, format: str = "markdown") -> str:
    """快速导出"""
    exporter = DataExporter()
    return exporter.export(results, filename, format)
